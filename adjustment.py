import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext
import re
import warnings
import os
import threading

# Supprime les warnings openpyxl sur les dates
warnings.filterwarnings('ignore', category=UserWarning)

# =====================================================
# APPLICATION CLASS
# =====================================================

class PortfolioProcessorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Portfolio Status Processor")
        self.root.geometry("900x800")
        self.root.configure(bg="#F5F5F5")
        
        # File paths
        self.base_file = tk.StringVar()
        self.daily_file = tk.StringVar()
        self.tv_file = tk.StringVar()
        self.output_folder = tk.StringVar()
        
        self.setup_ui()
        
    def setup_ui(self):
        """Setup the user interface"""
        
        # Main container
        main_container = tk.Frame(self.root, bg="#F5F5F5")
        main_container.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        # Header
        header = tk.Frame(main_container, bg="#F5F5F5")
        header.pack(fill=tk.X, pady=(0, 20))
        
        title = tk.Label(
            header,
            text="Portfolio Status Processor",
            font=("Segoe UI", 20, "bold"),
            bg="#F5F5F5",
            fg="#333333"
        )
        title.pack(anchor="w")
        
        subtitle = tk.Label(
            header,
            text="Select your files and run the processing",
            font=("Segoe UI", 10),
            bg="#F5F5F5",
            fg="#666666"
        )
        subtitle.pack(anchor="w")
        
        # Divider
        divider = tk.Frame(main_container, height=2, bg="#CCCCCC")
        divider.pack(fill=tk.X, pady=(0, 20))
        
        # File selection section
        self.create_file_section(main_container)
        
        # Process button
        button_frame = tk.Frame(main_container, bg="#F5F5F5")
        button_frame.pack(fill=tk.X, pady=20)
        
        process_btn = tk.Button(
            button_frame,
            text="Process Portfolio",
            command=self.process_files,
            font=("Segoe UI", 11, "bold"),
            bg="#0078D4",
            fg="white",
            padx=30,
            pady=12,
            relief=tk.FLAT,
            cursor="hand2"
        )
        process_btn.pack(side=tk.LEFT)
        
        # Status/Results section
        results_label = tk.Label(
            main_container,
            text="Processing Log:",
            font=("Segoe UI", 11, "bold"),
            bg="#F5F5F5",
            fg="#333333"
        )
        results_label.pack(anchor="w", pady=(10, 5))
        
        # Results text area
        self.results_text = scrolledtext.ScrolledText(
            main_container,
            height=15,
            width=100,
            font=("Segoe UI", 9),
            bg="white",
            fg="#333333",
            relief=tk.FLAT,
            borderwidth=1
        )
        self.results_text.pack(fill=tk.BOTH, expand=True, pady=(0, 10))
        
        # Status bar
        self.status_var = tk.StringVar(value="Ready")
        status_bar = tk.Label(
            main_container,
            textvariable=self.status_var,
            font=("Segoe UI", 9),
            bg="#E8E8E8",
            fg="#666666",
            padx=10,
            pady=8,
            relief=tk.FLAT,
            anchor="w"
        )
        status_bar.pack(fill=tk.X, side=tk.BOTTOM)
    
    def create_file_section(self, parent):
        """Create the file selection section"""
        
        file_frame = tk.Frame(parent, bg="white", relief=tk.FLAT, borderwidth=1)
        file_frame.pack(fill=tk.X, pady=(0, 20))
        
        inner_frame = tk.Frame(file_frame, bg="white")
        inner_frame.pack(fill=tk.BOTH, expand=True, padx=15, pady=15)
        
        self.create_file_picker(inner_frame, "Portfolio Base File:", self.base_file, self.browse_base_file, 0)
        self.create_file_picker(inner_frame, "Daily Sales Progress File:", self.daily_file, self.browse_daily_file, 1)
        self.create_file_picker(inner_frame, "Repart TV File:", self.tv_file, self.browse_tv_file, 2)
        self.create_folder_picker(inner_frame, "Output Folder:", self.output_folder, self.browse_output_folder, 3)
    
    def create_file_picker(self, parent, label_text, var, callback, row):
        """Create a file picker row"""
        
        label = tk.Label(parent, text=label_text, font=("Segoe UI", 10), bg="white", fg="#333333")
        label.grid(row=row, column=0, sticky="w", pady=8)
        
        path_frame = tk.Frame(parent, bg="#F9F9F9", relief=tk.SUNKEN, borderwidth=1)
        path_frame.grid(row=row, column=1, sticky="ew", padx=(10, 0), pady=8)
        
        path_label = tk.Label(path_frame, textvariable=var, font=("Segoe UI", 9), bg="#F9F9F9", fg="#666666", padx=10, pady=8, anchor="w")
        path_label.pack(fill=tk.BOTH, expand=True)
        
        btn = tk.Button(parent, text="Browse...", command=callback, font=("Segoe UI", 9), bg="#E8E8E8", fg="#333333", padx=15, pady=8, relief=tk.FLAT, cursor="hand2")
        btn.grid(row=row, column=2, sticky="ew", padx=(5, 0), pady=8)
        
        parent.columnconfigure(1, weight=1)
    
    def create_folder_picker(self, parent, label_text, var, callback, row):
        """Create a folder picker row"""
        
        label = tk.Label(parent, text=label_text, font=("Segoe UI", 10), bg="white", fg="#333333")
        label.grid(row=row, column=0, sticky="w", pady=8)
        
        path_frame = tk.Frame(parent, bg="#F9F9F9", relief=tk.SUNKEN, borderwidth=1)
        path_frame.grid(row=row, column=1, sticky="ew", padx=(10, 0), pady=8)
        
        path_label = tk.Label(path_frame, textvariable=var, font=("Segoe UI", 9), bg="#F9F9F9", fg="#666666", padx=10, pady=8, anchor="w")
        path_label.pack(fill=tk.BOTH, expand=True)
        
        btn = tk.Button(parent, text="Browse...", command=callback, font=("Segoe UI", 9), bg="#E8E8E8", fg="#333333", padx=15, pady=8, relief=tk.FLAT, cursor="hand2")
        btn.grid(row=row, column=2, sticky="ew", padx=(5, 0), pady=8)
        
        parent.columnconfigure(1, weight=1)
    
    def browse_base_file(self):
        file = filedialog.askopenfilename(title="Select Portfolio Base File", filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")])
        if file:
            self.base_file.set(file)
    
    def browse_daily_file(self):
        file = filedialog.askopenfilename(title="Select Daily Sales Progress File", filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")])
        if file:
            self.daily_file.set(file)
    
    def browse_tv_file(self):
        file = filedialog.askopenfilename(title="Select Repart TV File", filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")])
        if file:
            self.tv_file.set(file)
    
    def browse_output_folder(self):
        folder = filedialog.askdirectory(title="Select Output Folder")
        if folder:
            self.output_folder.set(folder)
    
    def log_message(self, message):
        self.results_text.insert(tk.END, message + "\n")
        self.results_text.see(tk.END)
        self.root.update()
    
    def clear_log(self):
        self.results_text.delete(1.0, tk.END)
    
    def validate_files(self):
        if not self.base_file.get():
            messagebox.showerror("Error", "Please select Portfolio Base File")
            return False
        if not self.daily_file.get():
            messagebox.showerror("Error", "Please select Daily Sales Progress File")
            return False
        if not self.tv_file.get():
            messagebox.showerror("Error", "Please select Repart TV File")
            return False
        if not self.output_folder.get():
            messagebox.showerror("Error", "Please select Output Folder")
            return False
        return True
    
    def process_files(self):
        if not self.validate_files():
            return
        thread = threading.Thread(target=self.run_processing)
        thread.start()
    
    def run_processing(self):
        """Run the actual processing"""
        self.clear_log()
        self.status_var.set("Processing...")
        
        try:
            BASE_FILE = self.base_file.get()
            DAILY_FILE = self.daily_file.get()
            TV_FILE = self.tv_file.get()
            OUTPUT_FOLDER = self.output_folder.get()
            OUTPUT_FILE = f"{OUTPUT_FOLDER}/Portfolio_Final.xlsx"
            
            self.log_message("Reading Portfolio Base file...")
            portfolio = pd.read_excel(BASE_FILE)
            portfolio.columns = portfolio.columns.astype(str).str.strip()
            
            self.log_message("Reading Daily Sales Progress file...")
            daily_df = pd.read_excel(DAILY_FILE, sheet_name="BASE")
            daily_df.columns = daily_df.columns.astype(str).str.strip()
            
            self.log_message("Reading Repart TV file...")
            raw = pd.read_excel(TV_FILE, sheet_name="REPART", header=None)
            raw.columns = raw.columns.astype(str).str.strip()
            
            alloc_tv = pd.read_excel(TV_FILE, sheet_name="REPART", header=4)
            alloc_tv.columns = alloc_tv.columns.astype(str).str.strip()
            
            # Print all Repart TV columns with their positions
            self.log_message("\n" + "="*60)
            self.log_message("REPART TV COLUMNS AND POSITIONS:")
            self.log_message("="*60)
            for idx, col in enumerate(alloc_tv.columns):
                self.log_message(f"Position {idx}: {col}")
            self.log_message("="*60 + "\n")
            
            # Detect weeks from RAW dataframe (merged header rows)
            self.log_message("\nDetecting weeks from merged header rows...")
            week_mapping = []
            
            # Scan raw dataframe for "SP W" patterns in header rows (0-4)
            for row_idx in range(min(5, len(raw))):
                self.log_message(f"\nScanning row {row_idx} of raw dataframe:")
                for col_idx in range(len(raw.columns)):
                    try:
                        cell_value = raw.iloc[row_idx, col_idx]
                        if pd.notna(cell_value):
                            cell_str = str(cell_value).strip()
                            if "SP W" in cell_str.upper():
                                # Found a week header in raw!
                                # The allocation is in the NEXT column
                                alloc_col_idx = col_idx + 1
                                if alloc_col_idx < len(alloc_tv.columns):
                                    alloc_col_name = alloc_tv.columns[alloc_col_idx]
                                    # Store BOTH the index and the label
                                    week_mapping.append((alloc_col_idx, cell_str))
                                    self.log_message(f"✓ Found week '{cell_str}' at raw col {col_idx} → allocation col index {alloc_col_idx} ('{alloc_col_name}')")
                    except:
                        pass
            
            self.log_message(f"\n═══════════════════════════════")
            self.log_message(f"Total weeks detected: {len(week_mapping)}")
            for alloc_col_idx, week_label in week_mapping:
                self.log_message(f"  {week_label} → allocations at column index {alloc_col_idx}")
            self.log_message(f"═══════════════════════════════\n")
            
            def parse_date(date_value):
                if pd.isna(date_value) or date_value == "":
                    return None
                try:
                    if hasattr(date_value, 'strftime'):
                        return date_value.strftime('%d/%m')
                    date_str = str(date_value).strip()
                    if not date_str or date_str.upper() in ['NAN', 'NONE', '']:
                        return None
                    try:
                        date_obj = pd.to_datetime(date_str, format='%d/%m/%Y')
                        return date_obj.strftime('%d/%m')
                    except ValueError:
                        pass
                    date_obj = pd.to_datetime(date_str)
                    return date_obj.strftime('%d/%m')
                except:
                    return None
            
            def search_daily_sales(model, customer_po, line_no):
                try:
                    matches = daily_df[
                        (daily_df["Item Name"].astype(str).str.strip() == str(model).strip()) &
                        (daily_df["Customer PO"].astype(str).str.strip() == str(customer_po).strip()) &
                        (daily_df["Line N°"].astype(str).str.strip() == str(line_no).strip())
                    ]
                    
                    if matches.empty:
                        return (None, None, None)
                    
                    row = matches.iloc[0]
                    
                    if "APPT DATE" in daily_df.columns:
                        appt_date_value = row["APPT DATE"]
                        formatted_date = parse_date(appt_date_value)
                        if formatted_date:
                            return (formatted_date, None, "Daily Sales - APPT DATE")
                    
                    if "Possible delivery date" in daily_df.columns:
                        delivery_value = row["Possible delivery date"]
                        formatted_date = parse_date(delivery_value)
                        if formatted_date:
                            return (formatted_date, None, "Daily Sales - Possible delivery")
                    
                    if "STATUT" in daily_df.columns:
                        statut = row["STATUT"]
                        if pd.notna(statut) and str(statut).strip():
                            return (str(statut).strip(), None, "Daily Sales - STATUT")
                    
                    return (None, None, None)
                except:
                    return (None, None, None)
            
            def search_repart_tv(line_no, order_no, is_back=False):
                try:
                    # Normalize inputs
                    order_no_str = str(order_no).strip() if order_no else ""
                    line_no_str = str(line_no).strip() if line_no else ""
                    
                    self.log_message(f"DEBUG: Searching Repart TV for Order N°={order_no_str}, Line N°={line_no_str}")
                    
                    if not order_no_str or not line_no_str:
                        self.log_message(f"DEBUG: Empty order or line number")
                        return (None, None, None)
                    
                    # Try exact match first
                    matches = alloc_tv[
                        (alloc_tv["Order N°"].astype(str).str.strip() == order_no_str) &
                        (alloc_tv["Line N°"].astype(str).str.strip() == line_no_str)
                    ]
                    
                    self.log_message(f"DEBUG: Found {len(matches)} exact matches")
                    
                    # If no exact match, try flexible matching with numeric conversion
                    if matches.empty:
                        self.log_message(f"DEBUG: Trying flexible matching...")
                        try:
                            order_no_numeric = int(float(order_no_str))
                            line_no_numeric = float(line_no_str)
                            
                            matches = alloc_tv[
                                (pd.to_numeric(alloc_tv["Order N°"].astype(str).str.strip(), errors='coerce') == order_no_numeric) &
                                (pd.to_numeric(alloc_tv["Line N°"].astype(str).str.strip(), errors='coerce') == line_no_numeric)
                            ]
                            self.log_message(f"DEBUG: Found {len(matches)} flexible matches")
                        except Exception as e:
                            self.log_message(f"DEBUG: Flexible matching error: {e}")
                    
                    if matches.empty:
                        self.log_message(f"DEBUG: No match found. Showing sample Repart TV data:")
                        for idx, tv_row in alloc_tv.head(3).iterrows():
                            try:
                                tv_order = str(tv_row.get("Order N°", "N/A")).strip()
                                tv_line = str(tv_row.get("Line N°", "N/A")).strip()
                                self.log_message(f"  Sample: Order N°={tv_order}, Line N°={tv_line}")
                            except:
                                pass
                        return (None, None, None)
                    
                    row = matches.iloc[0]
                    
                    # Get Q (Position 16), R (Position 17), T (Position 0 - STOCK DISPO)
                    try:
                        Q = float(row.iloc[16])  # QTE CDE (Customer asked)
                    except (ValueError, TypeError, IndexError):
                        Q = 0
                    
                    try:
                        R = float(row.iloc[17])  # QTE Restante (Remaining to deliver)
                    except (ValueError, TypeError, IndexError):
                        R = 0
                    
                    try:
                        T = float(row.iloc[0])  # STOCK DISPO (Available stock at position 0)
                    except (ValueError, TypeError):
                        T = 0
                    
                    self.log_message(f"DEBUG: Q={Q}, R={R}, T={T}")
                    
                    # Log allocation check
                    self.log_message(f"DEBUG: Checking {len(week_mapping)} allocations...")
                    for alloc_col_idx, week_label in week_mapping:
                        try:
                            qty = float(row.iloc[alloc_col_idx])
                            self.log_message(f"  {week_label} (col {alloc_col_idx}): {qty}")
                        except:
                            self.log_message(f"  {week_label} (col {alloc_col_idx}): ERROR reading value")
                    
                    # LOGIC 1: If T > 0
                    if T > 0:
                        return ("Disponibilité", None, "Repart TV - T > 0 (stock available)")
                    
                    # LOGIC 2: If R = 0 (everything already delivered)
                    if R == 0:
                        # First check if RDV column has a valid date (not just "-")
                        try:
                            rdv_value = row.get("RDV")
                            rdv_str = str(rdv_value).strip() if pd.notna(rdv_value) else ""
                            # Only return RDV if it's not empty and not just a dash
                            if rdv_str and rdv_str != "-":
                                rdv_date = parse_date(rdv_value)
                                if rdv_date:
                                    return (f"RDV {rdv_date}", "", "Repart TV - RDV date (R=0)")
                                else:
                                    return (f"RDV {rdv_str}", "", "Repart TV - RDV value (R=0)")
                        except:
                            pass
                        
                        # If no valid RDV, check for future allocations
                        for alloc_col_idx, week_label in week_mapping:
                            try:
                                qty = float(row.iloc[alloc_col_idx])
                                if qty > 0:
                                    # Extract week number from "SP W33" format
                                    match = re.search(r'SP\s*W(\d+)', str(week_label), re.IGNORECASE)
                                    if match:
                                        week_num = int(match.group(1))
                                        adjusted_week = f"S{week_num + 1}"
                                        return (f"Délai théorique {adjusted_week}", adjusted_week, "Repart TV - R=0 with allocation")
                            except (ValueError, TypeError):
                                pass
                        
                        return ("RDV à prendre", "", "Repart TV - R=0 (RDV to be scheduled)")
                    
                    # LOGIC 3: If Q == R
                    if Q == R:
                        return ("Pas de visibilité (pas encore traité)", None, "Repart TV - Q==R (not yet processed)")
                    
                    # LOGIC 4: If Q > R
                    if Q > R:
                        for alloc_col_idx, week_label in week_mapping:
                            try:
                                qty = float(row.iloc[alloc_col_idx])
                                if qty > 0:
                                    # Extract week number from "SP W33" format
                                    match = re.search(r'SP\s*W(\d+)', str(week_label), re.IGNORECASE)
                                    if match:
                                        week_num = int(match.group(1))
                                        adjusted_week = f"S{week_num + 1}"
                                        return (f"Délai théorique {adjusted_week}", adjusted_week, "Repart TV - Q>R with allocation")
                            except (ValueError, TypeError):
                                pass
                        return ("Pas de visibilité", None, "Repart TV - Q>R (no allocation)")
                    
                    return (None, None, None)
                except Exception as e:
                    self.log_message(f"DEBUG: Exception in search_repart_tv: {str(e)}")
                    return (None, None, None)
            
            def determine_line_status(row):
                model = row.get("Model")
                customer_po = row.get("Customer PO No.")
                line_no = row.get("Line No.")
                order_no = row.get("Order No.")
                division = str(row.get("Item Division", "")).strip()
                
                so_status = str(row.get("So Status(2)", "")).strip().lower()
                
                # PRIORITE 1: DELY STATUS
                if so_status == "dely":
                    return ("En cours de facturation", "", "So Status - Dely")
                
                # PRIORITE 2: PORTFOLIO APPOINTMENT DATE
                appointment = row.get("Appointment Date")
                if pd.notna(appointment):
                    formatted = parse_date(appointment)
                    if formatted:
                        return (f"RDV {formatted}", "", "Portfolio - Appointment Date")
                    else:
                        return (f"RDV {appointment}", "", "Portfolio - Appointment Date")
                
                # PRIORITE 3: FUTURE HOLD
                if str(row.get("Future Hold", "")).strip().upper() == "Y":
                    return ("Cadencement", "", "Portfolio - Future Hold")
                
                # PRIORITE 4: CREDIT HOLD
                if str(row.get("Credit Hold", "")).strip().upper() == "Y":
                    return ("Bloqué Credit", "", "Portfolio - Credit Hold")
                
                # PRIORITE 5: OVERDUE HOLD
                if str(row.get("Overdue Hold", "")).strip().upper() == "Y":
                    return ("Bloqué Credit", "", "Portfolio - Overdue Hold")
                
                # PRIORITE 6: CUSTOMER HOLD
                if str(row.get("Customer Hold", "")).strip().upper() == "Y":
                    return ("Bloqué Credit", "", "Portfolio - Customer Hold")
                
                # PRIORITE 7: SO STATUS = BACK
                if so_status == "back":
                    # Search for ALL products with back status, not just filtered by division
                    repart_status, repart_week, source = search_repart_tv(line_no, order_no, is_back=True)
                    if repart_status:
                        return (repart_status, repart_week if repart_week else "", source)
                    return ("Pas de visibilité", "", "No allocation found for Back status")
                
                # PRIORITE 8: SO STATUS = PICK
                elif so_status == "pick":
                    daily_status, daily_week, source = search_daily_sales(model, customer_po, line_no)
                    if daily_status:
                        return (daily_status, daily_week if daily_week else "", source)
                
                # PRIORITE 9: DAILY SALES PROGRESS
                daily_status, daily_week, source = search_daily_sales(model, customer_po, line_no)
                if daily_status:
                    return (daily_status, daily_week if daily_week else "", source)
                
                # PRIORITE 10: REPART TV
                # Search for ALL products, not just filtered by division
                repart_status, repart_week, source = search_repart_tv(line_no, order_no, is_back=False)
                if repart_status:
                    return (repart_status, repart_week if repart_week else "", source)
                
                # PRIORITE 11: PAS DE VISIBILITE
                return ("Pas de visibilité", "", "No match found")
            
            self.log_message("Processing lines...")
            results = portfolio.apply(determine_line_status, axis=1)
            
            portfolio["Portfolio Line Status"] = [x[0] for x in results]
            portfolio["Allocation Week"] = [x[1] for x in results]
            portfolio["Status Source"] = [x[2] for x in results]
            
            self.log_message(f"Exporting to {OUTPUT_FILE}...")
            portfolio.to_excel(OUTPUT_FILE, index=False)
            
            self.log_message("Applying formatting...")
            wb = load_workbook(OUTPUT_FILE)
            ws = wb.active
            
            yellow_fill = PatternFill(fill_type="solid", start_color="FFFF00", end_color="FFFF00")
            
            allocation_col = None
            for cell in ws[1]:
                if cell.value == "Allocation Week":
                    allocation_col = cell.column
                    break
            
            if allocation_col:
                for row_num in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row_num, column=allocation_col)
                    if cell.value and isinstance(cell.value, str):
                        if re.match(r'^S\d+$', cell.value.strip()):
                            cell.fill = yellow_fill
            
            wb.save(OUTPUT_FILE)
            
            self.log_message("\n" + "="*50)
            self.log_message("PROCESSING COMPLETED SUCCESSFULLY")
            self.log_message("="*50)
            self.log_message(f"Output file: {OUTPUT_FILE}")
            
            status_counts = portfolio["Portfolio Line Status"].value_counts()
            self.log_message("\nStatus Summary:")
            for status, count in status_counts.items():
                self.log_message(f"  {status}: {count}")
            
            self.status_var.set("Completed successfully")
            
            if messagebox.askyesno("Success", "Processing completed! Do you want to open the output file?"):
                os.startfile(OUTPUT_FILE)
        
        except Exception as e:
            self.log_message(f"\nERROR: {str(e)}")
            self.status_var.set("Error occurred")
            messagebox.showerror("Error", f"An error occurred:\n{str(e)}")

if __name__ == "__main__":
    root = tk.Tk()
    app = PortfolioProcessorApp(root)
    root.mainloop()